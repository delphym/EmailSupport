### Author: Daniel Mladek (skype: delphym) ###
import imaplib
import email
from email.header import decode_header
import pandas as pd
import matplotlib.pyplot as plt
from datetime import datetime
import ssl
import traceback
import configparser
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference

# Read email credentials from the config file
config = configparser.ConfigParser()
config.read('config.ini')

username = config.get('email', 'username')
password = config.get('email', 'password')
imap_server = config.get('email', 'imap_server')
imap_port = config.getint('email', 'imap_port')

# Define the start date for the query
start_date = "01-Apr-2024"

# Define filtering criteria
exclude_senders = [
    'testing@example.com'
    ,'lucido.leinteract@gmail.com'
    ,'mitaxebandilis@gmail.com'
]

include_subjects = [
    'TheJigsawPuzzles.com Feedback'
    ,'Feedback from TheSudoku.com'
    ,'Feedback - thesolitaire.com'
]

exclude_subjects = [
    'JIRA'
    ,'Undelivered Mail Returned to Sender'
]

# Define the folders to search
folders_to_search = [
    "INBOX", "Trash", "Spam",
    "INBOX.Examples", "INBOX.Personal",
    "INBOX.Resolved",
    "INBOX.Solitaire",
    "INBOX.Sudoku"
]

# Create a custom SSL context to allow smaller DH keys and bypass certificate verification
context = ssl.SSLContext(ssl.PROTOCOL_TLS_CLIENT)
context.set_ciphers('DEFAULT:@SECLEVEL=1')
context.check_hostname = False  # Disable hostname check
context.verify_mode = ssl.CERT_NONE  # Disable certificate verification

print("Connecting to the server...")
# Connect to the server using the custom SSL context and specified port
try:
    mail = imaplib.IMAP4_SSL(imap_server, port=imap_port, ssl_context=context)
    mail.login(username, password)
    print("Successfully connected to the server.")
except ssl.SSLError as e:
    print(f"SSL error: {e}")
    exit(1)
except imaplib.IMAP4.error as e:
    print(f"IMAP error: {e}")
    exit(1)

email_dates = []
folder_results = {}  # To store intermediate results for each folder

def list_folders():
    result, folders = mail.list()
    if result == 'OK':
        # Decode folder names from bytes to strings
        folder_paths = [folder.decode().split(' "." ')[-1].strip('"') for folder in folders]
        return folder_paths
    else:
        print("Failed to list folders.")
        return []

def search_emails(folder):
    print(f"Processing folder '{folder}'...")
    try:
        status, messages = mail.select(folder)
        if status != "OK":
            print(f"Error selecting folder '{folder}': {status}")
            return []
        # Use the start_date variable in the search query
        result, data = mail.search(None, f'(SINCE "{start_date}")')
        if result == "OK":
            return data[0].split()
        else:
            print(f"Failed to retrieve emails from folder '{folder}'.")
            return []
    except imaplib.IMAP4.error as e:
        print(f"Error selecting folder '{folder}': {e}")
        return []

def filter_emails(msg):
    # Get the sender
    sender = msg.get("From")
    if sender:
        sender = sender.lower()

    # Get the subject
    subject, encoding = decode_header(msg["Subject"])[0]
    if isinstance(subject, bytes):
        subject = subject.decode(encoding if encoding else "utf-8")

    # Apply filtering criteria
    if any(sender in sender_email for sender_email in exclude_senders):
        return False
    if not any(subject.startswith(inc_subj) for inc_subj in include_subjects):
        return False
    if any(exclude_subj in subject for exclude_subj in exclude_subjects):
        return False

    return True

available_folders = list_folders()
print(f"Available folders: {available_folders}")

for folder in folders_to_search:
    if folder in available_folders:
        email_ids = search_emails(folder)
        print(f"Found {len(email_ids)} emails in folder '{folder}'.")

        folder_dates = []

        for index, email_id in enumerate(email_ids):
            try:
                # Use BODY.PEEK to fetch the email without marking it as read
                res, msg_data = mail.fetch(email_id, "(BODY.PEEK[])")
                for response_part in msg_data:
                    if isinstance(response_part, tuple):
                        msg = email.message_from_bytes(response_part[1])

                        if filter_emails(msg):
                            date_tuple = email.utils.parsedate_tz(msg["Date"])
                            if date_tuple:
                                local_date = datetime.fromtimestamp(email.utils.mktime_tz(date_tuple))
                                folder_dates.append(local_date)
                                email_dates.append(local_date)

                # Log progress for every 100 emails processed
                if (index + 1) % 100 == 0:
                    print(f"Processed {index + 1} emails in folder '{folder}'.")

            except Exception as e:
                print(f"Error processing email ID {email_id} in folder '{folder}': {e}")
                print(traceback.format_exc())

        # Intermediate results for the current folder
        if folder_dates:
            df = pd.DataFrame(folder_dates, columns=["Date"])
            df["Date"] = pd.to_datetime(df["Date"])
            df.set_index("Date", inplace=True)
            email_counts = df.resample('D').size()
            folder_results[folder] = email_counts
            print(f"Intermediate results for folder '{folder}':")
            print(email_counts)
        else:
            print(f"No emails found in folder '{folder}'.")

    else:
        print(f"Folder '{folder}' not found on the server.")

# Close the connection
mail.logout()

print("Finished processing all folders.")

# Create a DataFrame from the dates for total aggregation
if email_dates:
    df = pd.DataFrame(email_dates, columns=["Date"])
    df["Date"] = pd.to_datetime(df["Date"])
    df.set_index("Date", inplace=True)

    # Count emails per day for total aggregation
    email_counts = df.resample('D').size()

    # Export to Excel
    excel_file = "email_analysis.xlsx"
    with pd.ExcelWriter(excel_file) as writer:
        # Write total counts to Excel
        total_counts_df = email_counts.reset_index()
        total_counts_df.columns = ["Date", "Total Emails"]
        total_counts_df.to_excel(writer, sheet_name="Total Emails", index=False)

        # Write individual folder counts to Excel
        for folder, counts in folder_results.items():
            folder_counts_df = counts.reset_index()
            folder_counts_df.columns = ["Date", "Emails"]
            folder_counts_df.to_excel(writer, sheet_name=folder.replace("/", "_"), index=False)

    print("Exported email counts to email_analysis.xlsx")

    # Add chart to the Excel file
    workbook = load_workbook(excel_file)
    sheet = workbook["Total Emails"]

    chart = LineChart()
    chart.title = "Number of Emails Received per Day"
    chart.style = 13
    chart.y_axis.title = "Number of Emails"
    chart.x_axis.title = "Date"

    data = Reference(sheet, min_col=2, min_row=1, max_col=2, max_row=sheet.max_row)
    categories = Reference(sheet, min_col=1, min_row=2, max_row=sheet.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    sheet.add_chart(chart, "E5")
    workbook.save(excel_file)

    print("Added chart to email_analysis.xlsx")

    # Plot the data using matplotlib (optional)
    plt.figure(figsize=(10, 6))
    email_counts.plot()
    plt.title("Number of Emails Received per Day")
    plt.xlabel("Date")
    plt.ylabel("Number of Emails")
    plt.grid(True)
    plt.show()
else:
    print("No emails found.")
